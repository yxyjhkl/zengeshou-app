# -*- coding: utf-8 -*-
"""
增额寿简版建议书生成助手 - 安卓版 Kivy UI
"""
import pdfplumber
import os
from datetime import datetime
from kivy.app import App
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.textinput import TextInput
from kivy.uix.slider import Slider
from kivy.uix.filechooser import FileChooserIconView
from kivy.uix.scrollview import ScrollView
from kivy.uix.gridlayout import GridLayout
from kivy.uix.popup import Popup
from kivy.core.window import Window
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

Window.softinput_mode = 'below_target'

class MainScreen(Screen):
    def __init__(self, **kwargs):
        super().__init__(**kwargs)
        self.data = []
        self.dividend_rate = 1.0
        
        layout = BoxLayout(orientation='vertical', padding=10, spacing=10)
        
        # 标题
        title = Label(
            text='[b][color=#e94560]增额寿简版建议书生成助手[/color][/b]',
            markup=True,
            font_size='20sp',
            size_hint_y=0.1
        )
        layout.add_widget(title)
        
        # 文件选择
        self.file_label = Label(
            text='未选择文件',
            color=(0.3, 0.8, 0.3, 1),
            size_hint_y=0.05
        )
        layout.add_widget(self.file_label)
        
        # 选择PDF按钮
        select_btn = Button(
            text='选择PDF文件',
            background_color=(0.1, 0.2, 0.4, 1),
            size_hint_y=0.08
        )
        select_btn.bind(on_press=self.select_pdf)
        layout.add_widget(select_btn)
        
        # 分红实现率
        rate_layout = BoxLayout(orientation='horizontal', size_hint_y=0.08)
        rate_label = Label(text='分红实现率:', font_size='14sp')
        self.rate_value = Label(text='1.0x (100%)', font_size='14sp', color=(0.9, 0.3, 0.3, 1))
        
        self.slider = Slider(min=0.5, max=2.0, value=1.0, step=0.1)
        self.slider.bind(value=self.on_slider_value)
        
        rate_layout.add_widget(rate_label)
        rate_layout.add_widget(self.rate_value)
        rate_layout.add_widget(self.slider)
        layout.add_widget(rate_layout)
        
        # 数据预览
        scroll = ScrollView(size_hint_y=0.5)
        self.data_layout = GridLayout(cols=9, size_hint_y=None, spacing=1)
        self.data_layout.bind(minimum_height=self.data_layout.setter('height'))
        
        # 添加表头
        for h in ['年度', '年龄', '累计保费', '身故总利益', '主险现价', '当年分红', '累积分红', '生存总利益', '演示增长率']:
            self.data_layout.add_widget(Label(text=h, size_hint_y=None, height=40, bold=True, color=(0.9, 0.9, 0.9, 1)))
        
        scroll.add_widget(self.data_layout)
        layout.add_widget(scroll)
        
        # 导出按钮
        export_btn = Button(
            text='导出Excel',
            background_color=(0.9, 0.3, 0.3, 1),
            size_hint_y=0.1
        )
        export_btn.bind(on_press=self.export_excel)
        layout.add_widget(export_btn)
        
        self.add_widget(layout)
    
    def select_pdf(self, instance):
        # 显示文件选择器弹窗
        filechooser = FileChooserIconView(path='/', filters=['*.pdf'])
        popup = Popup(title='选择PDF文件', content=filechooser, size_hint=(0.9, 0.9))
        
        def load_file(instance):
            if filechooser.selection:
                self.load_pdf(filechooser.selection[0])
                popup.dismiss()
        
        filechooser.bind(on_submit=lambda *args: load_file(None))
        popup.open()
    
    def load_pdf(self, path):
        try:
            self.data = process_pdf(path, self.dividend_rate)
            self.file_label.text = os.path.basename(path)
            self.update_data_display()
        except Exception as e:
            self.file_label.text = f'加载失败: {str(e)}'
    
    def on_slider_value(self, instance, value):
        self.dividend_rate = round(value, 1)
        self.rate_value.text = f'{self.dividend_rate}x ({int(self.dividend_rate*100)}%)'
        
        if self.data and self.file_label.text != '未选择文件':
            self.data = process_pdf(self.file_label.text, self.dividend_rate)
            self.update_data_display()
    
    def update_data_display(self):
        self.data_layout.clear_widgets()
        
        # 表头
        for h in ['年度', '年龄', '累计保费', '身故总利益', '主险现价', '当年分红', '累积分红', '生存总利益', '演示增长率']:
            self.data_layout.add_widget(Label(text=h, size_hint_y=None, height=40, bold=True))
        
        # 数据
        for row in self.data:
            for val in row:
                if isinstance(val, float) and val > 1:
                    txt = f'{val:,.0f}'
                elif isinstance(val, float):
                    txt = f'{val:.2f}%'
                else:
                    txt = str(val)
                self.data_layout.add_widget(Label(text=txt, size_hint_y=None, height=30))
    
    def export_excel(self, instance):
        if not self.data:
            popup = Popup(title='提示', content=Label(text='请先加载PDF文件'), size_hint=(0.6, 0.3))
            popup.open()
            return
        
        try:
            from kivy.storage.jsonstore import JsonStore
            from plyer import filechooser
            
            # 创建Excel
            wb = Workbook()
            ws = wb.active
            ws.title = '数据转换'
            
            hf = Font(bold=True, color='FFFFFF')
            hfill = PatternFill('solid', fgColor='4472C4')
            tb = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            
            headers = ['年度', '年龄', '累计保费', '身故总利益', '主险现价', '当年分红', '累积分红', '生存总利益', '演示增长率']
            for ci, h in enumerate(headers, 1):
                c = ws.cell(row=4, column=ci, value=h)
                c.font = hf
                c.fill = hfill
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border = tb
            
            for ri, row in enumerate(self.data, 5):
                for ci, v in enumerate(row, 1):
                    c = ws.cell(row=ri, column=ci, value=v)
                    c.border = tb
            
            for ri in range(5, 5 + len(self.data)):
                for ci in range(3, 8):
                    ws.cell(row=ri, column=ci).number_format = '#,##0'
                ws.cell(row=ri, column=9).number_format = '0.00'
            
            for col in 'ABCDEFGHI':
                ws.column_dimensions[col].width = 14
            
            # 保存到 Downloads
            output = '/sdcard/Download/增额寿建议书_转换.xlsx'
            wb.save(output)
            
            popup = Popup(title='成功', content=Label(text=f'已保存到:\n{output}'), size_hint=(0.8, 0.4))
            popup.open()
        except Exception as e:
            popup = Popup(title='错误', content=Label(text=str(e)), size_hint=(0.8, 0.4))
            popup.open()

def process_pdf(pdf_path, dividend_rate=1.0):
    """处理PDF文件"""
    all_data = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            for tbl in page.extract_tables():
                for row in tbl:
                    if row and row[0] and str(row[0]).isdigit() and len(row) >= 9:
                        all_data.append(row)
    
    if len(all_data) > 105:
        all_data = all_data[:105]
    
    raw_data = []
    for row in all_data:
        try:
            raw_data.append([
                int(row[0]) if row[0] else 0,
                int(row[1]) if row[1] else 0,
                int(float(row[3])) if row[3] else 0,
                int(float(row[6])) if row[6] else 0,
                int(float(row[4])) if row[4] else 0,
                int(float(row[7])) if row[7] else 0,
                int(float(row[8])) if row[8] else 0,
                0,
                float(row[5]) if row[5] and row[5] != '--' else 0
            ])
        except:
            continue
    
    # 应用分红实现率
    converted_data = []
    cumulative_dividend = 0
    for idx, row in enumerate(raw_data):
        new_row = row.copy()
        new_row[5] = int(row[5] * dividend_rate)
        cumulative_dividend += new_row[5]
        new_row[6] = cumulative_dividend
        new_row[7] = new_row[4] + new_row[6]
        
        if dividend_rate == 1.0:
            new_row[8] = row[8]
        else:
            if idx == 0:
                new_row[8] = row[8]
            else:
                prev_survival = converted_data[idx - 1][7]
                if prev_survival > 0:
                    new_row[8] = (new_row[7] - prev_survival) / prev_survival * 100
                else:
                    new_row[8] = 0
        
        converted_data.append(new_row)
    
    return converted_data

class MyApp(App):
    def build(self):
        sm = ScreenManager()
        sm.add_widget(MainScreen(name='main'))
        return sm

if __name__ == '__main__':
    MyApp().run()
